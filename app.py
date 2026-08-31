import os
import re
from dotenv import load_dotenv
from datetime import datetime
from io import BytesIO
from flask import Flask, render_template, redirect, url_for, flash, request, jsonify, send_file
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models import db, User, Product, Transaction

load_dotenv(".env")

app = Flask(__name__)

app.config['SECRET_KEY'] = os.getenv('SECRET_KEY')
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(app.instance_path, 'inventory.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

os.makedirs(app.instance_path, exist_ok=True)

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


@app.cli.command("init-db")
def init_db():
    db.create_all()

    if not User.query.first():
        supervisor = User(
            staff_id='s001',
            name='john',
            password_hash=generate_password_hash('john123'),
            role='supervisor',
            is_account_active=True
        )
        emp1 = User(
            staff_id='ib001',
            name='Razin',
            password_hash=generate_password_hash('razin123'),
            role='inbound_emp',
            is_account_active=True
        )
        emp2 = User(
            staff_id='ob001',
            name='Jeff',
            password_hash=generate_password_hash('jeff123'),
            role='outbound_emp',
            is_account_active=True
        )
        emp3 = User(
            staff_id='ob002',
            name='Hazrul',
            password_hash=generate_password_hash('hazrul123'),
            role='outbound_emp',
            is_account_active=True
        )
        emp4 = User(
            staff_id='ib002',
            name='Jack',
            password_hash=generate_password_hash('jack123'),
            role='inbound_emp',
            is_account_active=True
        )

        db.session.add_all([supervisor, emp1, emp2, emp3, emp4])

        products = [
            Product(sku='SKU-1001', name='Industrial Wooden Pallet', current_stock=50, min_threshold=10),
            Product(sku='SKU-1002', name='Heavy Duty Cardboard Box (L)', current_stock=200, min_threshold=30),
            Product(sku='SKU-1003', name='Bubble Wrap Roll 100m', current_stock=15, min_threshold=5),
            Product(sku='SKU-1004', name='Steel Warehouse Rack Shelving', current_stock=8, min_threshold=3),
        ]
        db.session.add_all(products)

        db.session.commit()
        print("Database initialized and seeded!")
    else:
        print("Database already contains data.")


# --- HELPER FUNCTIONS ---

def get_next_available_id_by_prefix(prefix):
    active_employees = User.query.filter(
        User.staff_id.like(f'{prefix}%'),
        User.is_account_active == True
    ).all()

    active_nums = set()
    for u in active_employees:
        match = re.search(r'\d+', u.staff_id)
        if match:
            active_nums.add(int(match.group()))

    candidate = 1
    while candidate in active_nums:
        candidate += 1

    return f"{prefix}{candidate:03d}"


def generate_next_sku():
    products = Product.query.filter(Product.sku.like('SKU-%')).all()
    max_num = 1000
    for p in products:
        match = re.search(r'\d+', p.sku)
        if match:
            num = int(match.group())
            if num > max_num:
                max_num = num
    return f"SKU-{max_num + 1}"


def is_using_default_password(user):
    """
    Checks whether a user's current password still matches the
    auto-generated default (first_name + '123') assigned at account
    creation. Works without a dedicated DB column by recomputing the
    expected default from the user's stored name and comparing hashes.
    """
    if not user or not user.name:
        return False
    first_name = user.name.split()[0].lower()
    expected_default = f"{first_name}123"
    return check_password_hash(user.password_hash, expected_default)


@app.context_processor
def inject_password_reminder():
    if current_user.is_authenticated:
        return dict(show_password_reminder=is_using_default_password(current_user))
    return dict(show_password_reminder=False)


# --- ROUTING LOGIC ---

@app.route('/')
def index():
    if current_user.is_authenticated:
        if current_user.role == 'supervisor':
            return redirect(url_for('supervisor_dashboard'))
        elif current_user.role == 'inbound_emp':
            return redirect(url_for('inbound_page'))
        elif current_user.role == 'outbound_emp':
            return redirect(url_for('outbound_page'))
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))

    if request.method == 'POST':
        staff_id = request.form.get('staff_id', '').strip().lower()
        password = request.form.get('password')

        user = User.query.filter_by(staff_id=staff_id).first()

        if not user or not user.is_account_active:
            flash('Invalid Staff ID or account deactivated.', 'danger')
        elif user.is_locked:
            flash('This account is locked after 3 failed login attempts. Please contact your supervisor for a password reset.', 'danger')
        elif check_password_hash(user.password_hash, password):
            user.failed_login_attempts = 0
            db.session.commit()
            login_user(user)
            flash(f'Logged in as {user.name} ({user.staff_id.upper()})', 'success')
            return redirect(url_for('index'))
        else:
            user.failed_login_attempts += 1
            if user.failed_login_attempts >= 3:
                user.is_locked = True
                user.lock_requested_at = datetime.utcnow()
                db.session.commit()
                flash('Account locked after 3 failed login attempts. A password reset request has been sent to your supervisor.', 'danger')
            else:
                db.session.commit()
                remaining = 3 - user.failed_login_attempts
                attempt_word = 'attempt' if remaining == 1 else 'attempts'
                flash(f'Invalid Staff ID or password. {remaining} {attempt_word} remaining before lockout.', 'danger')

    active_users = User.query.filter_by(is_account_active=True, is_locked=False).order_by(User.role, User.staff_id).all()

    return render_template('login.html', active_users=active_users)

# --- API AUDIT TRAIL ENDPOINTS ---

@app.route('/api/product-history/<int:product_id>')
@login_required
def product_history_api(product_id):
    if current_user.role not in ['inbound_emp', 'supervisor']:
        return jsonify({'error': 'Unauthorized'}), 403

    product = Product.query.get_or_404(product_id)
    txns = Transaction.query.filter_by(product_id=product.id, txn_type='INBOUND')\
        .order_by(Transaction.created_at.desc()).all()

    history_data = []
    for t in txns:
        actioned_at = None
        if t.approved_at:
            actioned_at = t.approved_at.strftime('%Y-%m-%d %H:%M:%S')
        elif t.rejected_at:
            actioned_at = t.rejected_at.strftime('%Y-%m-%d %H:%M:%S')

        history_data.append({
            'staff_id': t.requester.staff_id.upper(),
            'staff_name': t.requester_name_snapshot or t.requester.name,
            'quantity': t.quantity,
            'status': t.status,
            'date_time': t.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            'actioned_at': actioned_at,
            'actioned_by_staff_id': t.actioned_by.staff_id.upper() if t.actioned_by else None,
            'actioned_by_name': t.actioned_by_name_snapshot or (t.actioned_by.name if t.actioned_by else None),
            'rejection_reason': t.rejection_reason
        })

    return jsonify({
        'sku': product.sku,
        'name': product.name,
        'current_stock': product.current_stock,
        'history': history_data
    })


@app.route('/api/product-full-history/<int:product_id>')
@login_required
def product_full_history_api(product_id):
    if current_user.role not in ['inbound_emp', 'outbound_emp', 'supervisor']:
        return jsonify({'error': 'Unauthorized'}), 403

    product = Product.query.get_or_404(product_id)

    txns = Transaction.query.filter_by(product_id=product.id)\
        .order_by(Transaction.created_at.desc()).all()

    history_data = []
    for t in txns:
        actioned_at = None
        if t.approved_at:
            actioned_at = t.approved_at.strftime('%Y-%m-%d %H:%M:%S')
        elif t.rejected_at:
            actioned_at = t.rejected_at.strftime('%Y-%m-%d %H:%M:%S')

        history_data.append({
            'txn_type': t.txn_type,
            'staff_id': t.requester.staff_id.upper(),
            'staff_name': t.requester_name_snapshot or t.requester.name,
            'quantity': t.quantity,
            'status': t.status,
            'date_time': t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else None,
            'requested_at': t.created_at.strftime('%Y-%m-%d %H:%M:%S') if t.created_at else None,
            'approved_at': t.approved_at.strftime('%Y-%m-%d %H:%M:%S') if t.approved_at else None,
            'rejected_at': t.rejected_at.strftime('%Y-%m-%d %H:%M:%S') if t.rejected_at else None,
            'actioned_at': actioned_at,
            'actioned_by_staff_id': t.actioned_by.staff_id.upper() if t.actioned_by else None,
            'actioned_by_name': t.actioned_by_name_snapshot or (t.actioned_by.name if t.actioned_by else None),
            'rejection_reason': t.rejection_reason
        })

    return jsonify({
        'sku': product.sku,
        'name': product.name,
        'current_stock': product.current_stock,
        'history': history_data
    })


# --- INBOUND ROUTES ---

@app.route('/inbound')
@login_required
def inbound_page():
    if current_user.role not in ['inbound_emp', 'supervisor']:
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('index'))

    products = Product.query.order_by(Product.sku.asc()).all()
    inbound_history = Transaction.query.filter_by(txn_type='INBOUND')\
        .order_by(Transaction.created_at.desc()).all()
    user_transactions = Transaction.query.filter_by(user_id=current_user.id, txn_type='INBOUND')\
        .order_by(Transaction.created_at.desc()).all()

    return render_template('inbound.html', products=products, inbound_history=inbound_history, user_transactions=user_transactions)


@app.route('/submit-inbound-existing', methods=['POST'])
@login_required
def submit_inbound_existing():
    if current_user.role not in ['inbound_emp', 'supervisor']:
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    product_id = request.form.get('product_id')
    quantity = int(request.form.get('quantity', 0))

    if not product_id or quantity <= 0:
        flash('Please select a product and enter a valid quantity.', 'warning')
        return redirect(url_for('inbound_page'))

    txn = Transaction(
        txn_type='INBOUND',
        quantity=quantity,
        status='PENDING',
        user_id=current_user.id,
        product_id=product_id,
        requester_name_snapshot=current_user.name
    )
    db.session.add(txn)
    db.session.commit()

    flash(f'Inbound request submitted for approval (+{quantity} units)!', 'success')
    return redirect(url_for('inbound_page'))


@app.route('/submit-inbound-new-batch', methods=['POST'])
@login_required
def submit_inbound_new_batch():
    if current_user.role not in ['inbound_emp', 'supervisor']:
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    names = request.form.getlist('new_item_name[]')
    quantities = request.form.getlist('new_item_qty[]')

    items_created = 0

    for name, qty_str in zip(names, quantities):
        item_name = name.strip()
        if item_name and qty_str.isdigit() and int(qty_str) > 0:
            qty = int(qty_str)

            new_sku = generate_next_sku()

            new_product = Product(
                sku=new_sku,
                name=item_name,
                current_stock=0,
                min_threshold=5
            )
            db.session.add(new_product)
            db.session.flush()

            txn = Transaction(
                txn_type='INBOUND',
                quantity=qty,
                status='PENDING',
                user_id=current_user.id,
                product_id=new_product.id,
                requester_name_snapshot=current_user.name
            )
            db.session.add(txn)
            items_created += 1

    db.session.commit()

    if items_created > 0:
        flash(f'Successfully registered {items_created} new item(s) with autogenerated SKUs and submitted inbound requests!', 'success')
    else:
        flash('No valid items were submitted.', 'warning')

    return redirect(url_for('inbound_page'))


# --- OUTBOUND ROUTES ---

@app.route('/outbound')
@login_required
def outbound_page():
    if current_user.role not in ['outbound_emp', 'supervisor']:
        flash('Unauthorized access.', 'danger')
        return redirect(url_for('index'))

    products = Product.query.order_by(Product.sku.asc()).all()
    user_transactions = Transaction.query.filter_by(user_id=current_user.id)\
        .order_by(Transaction.created_at.desc()).all()

    return render_template('outbound.html', products=products, user_transactions=user_transactions)


@app.route('/submit-outbound', methods=['POST'])
@login_required
def submit_outbound():
    if current_user.role not in ['outbound_emp', 'supervisor']:
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    product_id = request.form.get('product_id')
    quantity_str = request.form.get('quantity', '0')

    if not product_id or not quantity_str.isdigit() or int(quantity_str) <= 0:
        flash('Please select a product and enter a valid quantity.', 'warning')
        return redirect(url_for('outbound_page'))

    quantity = int(quantity_str)
    product = Product.query.get(product_id)

    if not product:
        flash('Selected product does not exist.', 'danger')
        return redirect(url_for('outbound_page'))

    if quantity > product.current_stock:
        flash(f'Cannot withdraw {quantity} units. Only {product.current_stock} currently available in stock!', 'warning')
        return redirect(url_for('outbound_page'))

    txn = Transaction(
        txn_type='OUTBOUND',
        quantity=quantity,
        status='PENDING',
        user_id=current_user.id,
        product_id=product_id,
        requester_name_snapshot=current_user.name
    )
    db.session.add(txn)
    db.session.commit()

    flash(f'Outbound request for {quantity} units of {product.sku} submitted for supervisor approval!', 'success')
    return redirect(url_for('outbound_page'))


# --- SUPERVISOR ROUTES ---

@app.route('/supervisor')
@login_required
def supervisor_dashboard():
    if current_user.role != 'supervisor':
        flash('Supervisor access required.', 'danger')
        return redirect(url_for('index'))

    pending_txns = Transaction.query.filter_by(status='PENDING').order_by(Transaction.created_at.asc()).all()
    products = Product.query.all()
    employees = User.query.order_by(User.staff_id.asc()).all()
    locked_users = User.query.filter_by(is_locked=True).order_by(User.lock_requested_at.asc()).all()

    inbound_count = User.query.filter_by(role='inbound_emp', is_account_active=True).count()
    outbound_count = User.query.filter_by(role='outbound_emp', is_account_active=True).count()

    forced_role = None
    if inbound_count < outbound_count:
        forced_role = 'inbound_emp'
    elif outbound_count < inbound_count:
        forced_role = 'outbound_emp'

    next_ib_id = get_next_available_id_by_prefix('ib')
    next_ob_id = get_next_available_id_by_prefix('ob')
    next_s_id = get_next_available_id_by_prefix('s')

    return render_template(
        'supervisor.html',
        pending_txns=pending_txns,
        products=products,
        employees=employees,
        locked_users=locked_users,
        inbound_count=inbound_count,
        outbound_count=outbound_count,
        forced_role=forced_role,
        next_ib_id=next_ib_id,
        next_ob_id=next_ob_id,
        next_s_id=next_s_id
    )


@app.route('/supervisor/report')
@login_required
def generate_stock_report():
    if current_user.role != 'supervisor':
        flash('Access denied.', 'danger')
        return redirect(url_for('login'))

    wb = Workbook()

    # ---- Sheet 1: Stock Summary ----
    summary = wb.active
    summary.title = 'Stock Summary'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    low_stock_fill = PatternFill(start_color='F8CBAD', end_color='F8CBAD', fill_type='solid')

    summary_headers = ['SKU', 'Product Name', 'Current Stock', 'Safety Threshold', 'Status']
    summary.append(summary_headers)
    for col_num, _ in enumerate(summary_headers, 1):
        cell = summary.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    products = Product.query.order_by(Product.sku).all()
    for product in products:
        is_low = product.current_stock < product.min_threshold
        status = 'LOW STOCK' if is_low else 'OK'
        row = [product.sku, product.name, product.current_stock, product.min_threshold, status]
        summary.append(row)
        if is_low:
            row_num = summary.max_row
            for col_num in range(1, 6):
                summary.cell(row=row_num, column=col_num).fill = low_stock_fill

    for col_num, header in enumerate(summary_headers, 1):
        summary.column_dimensions[chr(64 + col_num)].width = max(len(header) + 4, 18)

    # ---- Sheet 2: Full Audit Trail ----
    audit = wb.create_sheet('Full Audit Trail')

    audit_headers = ['Txn ID', 'Type', 'SKU', 'Product Name', 'Quantity', 'Status',
                      'Requester', 'Actioned By', 'Created At', 'Approved At',
                      'Rejected At', 'Rejection Reason']
    audit.append(audit_headers)
    for col_num, _ in enumerate(audit_headers, 1):
        cell = audit.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    transactions = Transaction.query.order_by(Transaction.created_at).all()
    for txn in transactions:
        requester_name = txn.requester_name_snapshot or (txn.requester.name if txn.requester else 'Unknown')
        actioned_by_name = txn.actioned_by_name_snapshot or (txn.actioned_by.name if txn.actioned_by else '')
        audit.append([
            txn.id,
            txn.txn_type,
            txn.product.sku if txn.product else '',
            txn.product.name if txn.product else '',
            txn.quantity,
            txn.status,
            requester_name,
            actioned_by_name,
            txn.created_at.strftime('%Y-%m-%d %H:%M:%S') if txn.created_at else '',
            txn.approved_at.strftime('%Y-%m-%d %H:%M:%S') if txn.approved_at else '',
            txn.rejected_at.strftime('%Y-%m-%d %H:%M:%S') if txn.rejected_at else '',
            txn.rejection_reason or '',
        ])

    for col_num, header in enumerate(audit_headers, 1):
        audit.column_dimensions[chr(64 + col_num)].width = max(len(header) + 4, 16)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"stock_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/supervisor/add-employee', methods=['POST'])
@login_required
def add_employee():
    if current_user.role != 'supervisor':
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    first_name = request.form.get('first_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    role = request.form.get('role')

    if not first_name or not last_name:
        flash('First and last name are required.', 'warning')
        return redirect(url_for('supervisor_dashboard'))

    name = f"{first_name.capitalize()} {last_name.capitalize()}"

    if role != 'supervisor':
        inbound_count = User.query.filter_by(role='inbound_emp', is_account_active=True).count()
        outbound_count = User.query.filter_by(role='outbound_emp', is_account_active=True).count()

        if inbound_count < outbound_count:
            role = 'inbound_emp'
        elif outbound_count < inbound_count:
            role = 'outbound_emp'

    if role == 'supervisor':
        prefix = 's'
    elif role == 'inbound_emp':
        prefix = 'ib'
    else:
        prefix = 'ob'

    new_staff_id = get_next_available_id_by_prefix(prefix)

    auto_password = f"{first_name.lower()}123"

    existing_user = User.query.filter_by(staff_id=new_staff_id).first()
    if existing_user:
        existing_user.name = name
        existing_user.password_hash = generate_password_hash(auto_password)
        existing_user.role = role
        existing_user.is_account_active = True
    else:
        new_user = User(
            staff_id=new_staff_id,
            name=name,
            password_hash=generate_password_hash(auto_password),
            role=role,
            is_account_active=True
        )
        db.session.add(new_user)

    db.session.commit()

    flash(f'New Employee Registered! Staff ID: {new_staff_id} | Name: {name} | Initial Password: {auto_password}', 'success')
    return redirect(url_for('supervisor_dashboard'))


@app.route('/supervisor/toggle-employee/<int:user_id>')
@login_required
def toggle_employee_status(user_id):
    if current_user.role != 'supervisor':
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)
    if user.id == current_user.id:
        flash('You cannot deactivate your own account.', 'danger')
        return redirect(url_for('supervisor_dashboard'))

    user.is_account_active = not user.is_account_active
    db.session.commit()

    status_str = "deactivated (Staff ID unassigned for recycling)" if not user.is_account_active else "reactivated"
    flash(f'Account for {user.name} ({user.staff_id}) has been {status_str}.', 'info')
    return redirect(url_for('supervisor_dashboard'))


@app.route('/supervisor/edit-employee/<int:user_id>', methods=['POST'])
@login_required
def edit_employee(user_id):
    if current_user.role != 'supervisor':
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)

    first_name = request.form.get('edit_first_name', '').strip()
    last_name = request.form.get('edit_last_name', '').strip()

    if not first_name:
        flash('First name is required.', 'warning')
        return redirect(url_for('supervisor_dashboard'))

    old_name = user.name
    if last_name:
        new_name = f"{first_name.capitalize()} {last_name.capitalize()}"
    else:
        new_name = first_name.capitalize()

    user.name = new_name
    db.session.commit()

    flash(f'Updated name for {user.staff_id}: "{old_name}" -> "{new_name}".', 'success')
    return redirect(url_for('supervisor_dashboard'))


@app.route('/supervisor/reset-password/<int:user_id>')
@login_required
def reset_employee_password(user_id):
    if current_user.role != 'supervisor':
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    user = User.query.get_or_404(user_id)

    first_name = user.name.split()[0].lower()
    default_password = f"{first_name}123"

    user.password_hash = generate_password_hash(default_password)
    user.is_locked = False
    user.failed_login_attempts = 0
    user.lock_requested_at = None
    db.session.commit()

    flash(f'Password for {user.name} ({user.staff_id}) has been reset to the default password ({default_password}).', 'success')
    return redirect(url_for('supervisor_dashboard'))


@app.route('/supervisor/approve/<int:txn_id>')
@login_required
def approve_transaction(txn_id):
    if current_user.role != 'supervisor':
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    txn = Transaction.query.get_or_404(txn_id)
    if txn.status == 'PENDING':
        product = Product.query.get(txn.product_id)

        if txn.txn_type == 'INBOUND':
            product.current_stock += txn.quantity
        elif txn.txn_type == 'OUTBOUND':
            if product.current_stock >= txn.quantity:
                product.current_stock -= txn.quantity
            else:
                flash('Approval failed: Insufficient physical stock available.', 'danger')
                return redirect(url_for('supervisor_dashboard'))

        txn.status = 'APPROVED'
        txn.approved_at = datetime.utcnow()
        txn.actioned_by_user_id = current_user.id
        txn.actioned_by_name_snapshot = current_user.name
        db.session.commit()
        flash(f'Transaction #{txn.id} approved! Inventory updated.', 'success')

    return redirect(url_for('supervisor_dashboard'))


@app.route('/supervisor/reject/<int:txn_id>', methods=['POST'])
@login_required
def reject_transaction(txn_id):
    if current_user.role != 'supervisor':
        flash('Unauthorized.', 'danger')
        return redirect(url_for('index'))

    txn = Transaction.query.get_or_404(txn_id)
    if txn.status == 'PENDING':
        reason = request.form.get('rejection_reason')
        txn.status = 'REJECTED'
        txn.rejection_reason = reason
        txn.rejected_at = datetime.utcnow()
        txn.actioned_by_user_id = current_user.id
        txn.actioned_by_name_snapshot = current_user.name
        db.session.commit()
        flash(f'Transaction #{txn.id} was rejected.', 'info')

    return redirect(url_for('supervisor_dashboard'))


@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')

        if not check_password_hash(current_user.password_hash, old_password):
            flash('Current password is incorrect.', 'danger')
            return redirect(url_for('change_password'))

        if not new_password or not confirm_password:
            flash('Please fill in both new password fields.', 'warning')
            return redirect(url_for('change_password'))

        if new_password != confirm_password:
            flash('New password and confirmation do not match.', 'warning')
            return redirect(url_for('change_password'))

        if new_password == old_password:
            flash('New password must be different from your current password.', 'warning')
            return redirect(url_for('change_password'))

        current_user.password_hash = generate_password_hash(new_password)
        db.session.commit()

        flash('Password updated successfully!', 'success')
        return redirect(url_for('index'))

    return render_template('change_password.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully.', 'info')
    return redirect(url_for('login'))


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
